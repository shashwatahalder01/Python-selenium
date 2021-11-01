import openpyxl

from openpyxl.styles import Alignment
from datetime import date
import pathlib
import sys
import os

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno, colno).value


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


# def writecol(file, sheetName, col, data):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     row = 1
#     for d in data:
#         sheet.cell(row, col).value = d
#         workbook.save(file)
#         row += 1


def readsinglerow(file, sheetName, rownum, colnum):
    single_row = []
    for i in range(1, colnum + 1):
        data = readData(file, sheetName, rownum, i)
        single_row.append(data)
    return single_row


def readsinglecol(file, sheetName, rownum, colnum):
    single_col = []
    for i in range(1, rownum + 1):
        data = readData(file, sheetName, i, colnum)
        single_col.append(data)
    return single_col


def writelistoflist(file, sheetName, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    row = 1
    for element in data:
        col = 1
        for d in element:
            if d:
                sheet.cell(row, col).value = d
            col += 1
        row += 1
    wb.save(file)


def writesinglerow(file, sheetName, rownum, colnum, scol, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    for i in range(1, colnum + 1):
        sheet.cell(rownum, i + scol).value = data[i - 1]
        # print(data[i - 1])
    wb.save(file)

def writesinglerowlink(file, sheetName, rownum, colnum, scol, data, colnumtoaddlink, hyperlink):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    for i in range(1, colnum + 1):
        sheet.cell(rownum, i + scol).value = data[i-1]
        if i == colnumtoaddlink:
            sheet.cell(rownum, i + scol).hyperlink = hyperlink[i-1]
            sheet.cell(rownum, i + scol).value = data[i-1]
            sheet.cell(rownum, i + scol).style = "Hyperlink"
        # print(data[i - 1])
    wb.save(file)


def mergecell(file, sheetName, mcell1, mcell2, rownum, colnum, ):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.merge_cells(f'{mcell1}:{mcell2}')
    cell = sheet.cell(rownum, colnum)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(file)


def writeinmergecell(file, sheetName, mcell1, mcell2, rownum, colnum, scol, data):
    mergecell(file, sheetName, mcell1, mcell2, rownum, scol + 1)
    writesinglerow(file, sheetName, rownum, colnum, scol, data)


def createFile(filename, sheetname):
    wb = openpyxl.Workbook()
    wb.create_sheet(sheetname, 0)
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(filename=filename)



# def createlktestresultExcelfile():
#     wb = openpyxl.Workbook()
#     # ws = wb.active
#     wb.create_sheet("GetAppointments", 0)
#     wb.create_sheet("GetMedications", 1)
#     wb.create_sheet("GetAllergies", 2)
#     wb.create_sheet("GetImmunizations", 3)
#     wb.create_sheet("GetLabResults", 4)
#     sheet = wb['Sheet']
#     wb.remove(sheet)
#     wb.save(filename=lktestresultExcelfile)

# path = pathlib.Path(__file__).parent.parent / "Data/ExcelData/LkTestResult.xlsx"
# data = ['header']
# writeinmergecell(path, 'a', 'A1', 'C1', 1, 1, 0, data)
# mergecell(path, 'a', 1, 3, 0, "data")
# writesinglerow(path, 'a', 1, 1, 0, data)

def insert_col(file, sheet_name, col_num):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    sheet.insert_cols(col_num)
    wb.save(file)
    
def get_sheet_names(file):
    wb = openpyxl.load_workbook(file)
    sheets = wb.sheetnames
    print(sheets)


def delete_all_sheets(file):
    wb = openpyxl.load_workbook(file)
    sheets = wb.sheetnames
    for sheet in sheets:
        wb.remove(wb[sheet])
    wb.create_sheet("Sheet", 0)  # need at-least one sheet
    wb.save(file)


def create_sheet(file, sheet_name, sheet_index):
    wb = openpyxl.Workbook()
    wb.create_sheet(sheet_name, sheet_index)
    wb.save(file)

